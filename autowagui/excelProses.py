import os
import openpyxl
from openpyxl import load_workbook, Workbook
from pyautogui import click, typewrite, press, hotkey, alert, confirm, prompt, password


# Open the file in read mode and read the contents of the file into a variable
# excel operations

# location of the file
path = "C:\\Users\\Muhammad Husni\\Documents\\RPA\\Book1.xlsx"

wb = openpyxl.load_workbook(path)
sheet = wb.active
cell = sheet.cell(row=1, column=1)
a = sheet.cell(row=2, column=2)

os.system('start "chrome" "https://api.whatsapp.com//send?phone=+62'+str(a.value)+'&text='+cell.value+'"')

